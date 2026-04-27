import streamlit as st
import cv2
import numpy as np
from collections import defaultdict
import pandas as pd
from datetime import datetime
import json
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import sqlite3
import os
import pickle
from enum import Enum

# ==================== Configuration de la page ====================
st.set_page_config(
    page_title="Gestionnaire d'Inventaire Multi-Pièces",
    page_icon="📦",
    layout="wide"
)

# ==================== Algorithmes de détection ====================

class AlgorithmeDetection(Enum):
    WATERSHED = "watershed"
    HOUGH_CERCLES = "hough_cercles"
    CONTOURS = "contours"

def detecter_pieces_avancee(image, algorithme="watershed", params=None):
    """
    Détecte et compte les pièces dans une image avec différents algorithmes.
    
    Args:
        image: image BGR
        algorithme: "watershed", "hough_cercles", ou "contours"
        params: dictionnaire de paramètres spécifiques
    
    Returns:
        (image_annotee, nb_pieces, debug_info)
    """
    resultat = image.copy()
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Paramètres par défaut
    if params is None:
        params = {}
    
    if algorithme == "watershed":
        return detecter_avec_watershed(image, gray, resultat, params)
    elif algorithme == "hough_cercles":
        return detecter_avec_hough_cercles(image, gray, resultat, params)
    else:
        return detecter_avec_contours(image, gray, resultat, params)

def detecter_avec_watershed(image, gray, resultat, params):
    """Détection par watershed - idéal pour objets qui se touchent"""
    
    # Paramètres
    seuil_distance = params.get('seuil_distance', 0.5)
    kernel_size = params.get('kernel_size', 3)
    aire_min = params.get('aire_min', 200)
    aire_max = params.get('aire_max', 10000)
    
    # 1. Réduction du bruit
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # 2. Binarisation adaptative
    _, binary = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # 3. Opérations morphologiques
    kernel = np.ones((kernel_size, kernel_size), np.uint8)
    opening = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=2)
    
    # 4. Arrière-plan sûr
    sure_bg = cv2.dilate(opening, kernel, iterations=3)
    
    # 5. Distance Transform
    dist_transform = cv2.distanceTransform(opening, cv2.DIST_L2, 5)
    _, sure_fg = cv2.threshold(dist_transform, seuil_distance * dist_transform.max(), 255, 0)
    sure_fg = np.uint8(sure_fg)
    
    # 6. Régions inconnues
    unknown = cv2.subtract(sure_bg, sure_fg)
    
    # 7. Marquage
    _, markers = cv2.connectedComponents(sure_fg)
    markers = markers + 1
    markers[unknown == 255] = 0
    
    # 8. Watershed
    markers = cv2.watershed(image, markers)
    
    # 9. Filtrage
    compteur_pieces = 0
    
    for marker_id in range(2, markers.max() + 1):
        mask = np.uint8(markers == marker_id)
        aire = cv2.countNonZero(mask)
        
        if aire_min < aire < aire_max:
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if contours:
                compteur_pieces += 1
                cv2.drawContours(resultat, [contours[0]], -1, (0, 255, 0), 2)
                
                M = cv2.moments(contours[0])
                if M["m00"] != 0:
                    cx = int(M["m10"] / M["m00"])
                    cy = int(M["m01"] / M["m00"])
                    cv2.circle(resultat, (cx, cy), 4, (0, 0, 255), -1)
                    cv2.putText(resultat, str(compteur_pieces), (cx - 10, cy - 10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
    
    # Frontières watershed en rouge
    resultat[markers == -1] = [0, 0, 255]
    
    debug_info = {
        'algorithm': 'watershed',
        'n_objects': compteur_pieces,
        'markers_found': markers.max()
    }
    
    return resultat, compteur_pieces, debug_info

def detecter_avec_hough_cercles(image, gray, resultat, params):
    """Détection par transformée de Hough pour cercles"""
    
    # Paramètres
    dp = params.get('dp', 1.2)
    min_dist = params.get('min_dist', 20)
    param1 = params.get('param1', 100)
    param2 = params.get('param2', 30)
    min_radius = params.get('min_radius', 10)
    max_radius = params.get('max_radius', 100)
    
    # Réduction du bruit
    blur = cv2.medianBlur(gray, 5)
    
    # Détection des cercles
    cercles = cv2.HoughCircles(
        blur,
        cv2.HOUGH_GRADIENT,
        dp=dp,
        minDist=min_dist,
        param1=param1,
        param2=param2,
        minRadius=min_radius,
        maxRadius=max_radius
    )
    
    compteur_pieces = 0
    
    if cercles is not None:
        cercles = np.round(cercles[0, :]).astype("int")
        
        for (x, y, r) in cercles:
            h, w = image.shape[:2]
            if x - r > 0 and x + r < w and y - r > 0 and y + r < h:
                cv2.circle(resultat, (x, y), r, (0, 255, 0), 2)
                cv2.circle(resultat, (x, y), 3, (0, 0, 255), -1)
                
                compteur_pieces += 1
                cv2.putText(resultat, str(compteur_pieces), (x - 15, y - 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
    
    debug_info = {
        'algorithm': 'hough_cercles',
        'n_objects': compteur_pieces,
        'circles_found': len(cercles) if cercles is not None else 0
    }
    
    return resultat, compteur_pieces, debug_info

def detecter_avec_contours(image, gray, resultat, params):
    """Détection classique par contours"""
    
    # Paramètres
    seuil_canny1 = params.get('seuil_canny1', 50)
    seuil_canny2 = params.get('seuil_canny2', 150)
    aire_min = params.get('aire_min', 200)
    kernel_size = params.get('kernel_size', 3)
    
    # Flou
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # Canny
    edges = cv2.Canny(blur, seuil_canny1, seuil_canny2)
    
    # Morphologie
    kernel = np.ones((kernel_size, kernel_size), np.uint8)
    edges = cv2.dilate(edges, kernel, iterations=2)
    edges = cv2.erode(edges, kernel, iterations=1)
    
    # Contours
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Filtrage
    pieces_valides = []
    for contour in contours:
        aire = cv2.contourArea(contour)
        if aire > aire_min:
            pieces_valides.append(contour)
    
    compteur_pieces = len(pieces_valides)
    
    # Dessin
    for i, contour in enumerate(pieces_valides):
        cv2.drawContours(resultat, [contour], -1, (0, 255, 0), 2)
        
        M = cv2.moments(contour)
        if M["m00"] != 0:
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])
            cv2.circle(resultat, (cx, cy), 3, (0, 0, 255), -1)
            cv2.putText(resultat, str(i + 1), (cx - 10, cy - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
    
    debug_info = {
        'algorithm': 'contours',
        'n_objects': compteur_pieces,
        'contours_found': len(contours)
    }
    
    return resultat, compteur_pieces, debug_info

def detecter_pieces(image, algorithme="watershed"):
    """Version de compatibilité pour l'interface"""
    resultat, nb, _ = detecter_pieces_avancee(image, algorithme)
    
    # Ajouter le compteur
    cv2.putText(resultat, f"Pieces: {nb} ({algorithme})", (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
    
    return resultat, nb

# ==================== Fonctions de persistance SQLite ====================

def init_database():
    """Initialise la base de données SQLite"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    
    # Table des articles
    c.execute('''CREATE TABLE IF NOT EXISTS articles
                 (code TEXT PRIMARY KEY,
                  libelle TEXT,
                  emplacement TEXT,
                  date_creation TEXT)''')
    
    # Table des photos
    c.execute('''CREATE TABLE IF NOT EXISTS photos
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  code_article TEXT,
                  timestamp TEXT,
                  nb_pieces INTEGER,
                  algorithme TEXT,
                  image_originale TEXT,
                  image_analyse TEXT,
                  FOREIGN KEY (code_article) REFERENCES articles(code))''')
    
    conn.commit()
    conn.close()

def charger_donnees():
    """Charge les données depuis SQLite"""
    gestionnaire = GestionnairePieces()
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    
    # Charger les articles
    c.execute("SELECT code, libelle, emplacement, date_creation FROM articles")
    articles = c.fetchall()
    
    for code, libelle, emplacement, date_creation in articles:
        gestionnaire.articles[code] = {
            'libelle': libelle,
            'photos': [],
            'emplacement': emplacement,
            'date_creation': date_creation
        }
    
    # Charger les photos
    c.execute("SELECT code_article, timestamp, nb_pieces, algorithme, image_originale, image_analyse, id FROM photos ORDER BY timestamp")
    photos = c.fetchall()
    
    for code_article, timestamp, nb_pieces, algorithme, img_originale, img_analyse, photo_id in photos:
        if code_article in gestionnaire.articles:
            photo_data = {
                'timestamp': timestamp,
                'nb_pieces': nb_pieces,
                'algorithme': algorithme,
                'image_originale': img_originale,
                'image_analyse': img_analyse,
                'id': len(gestionnaire.articles[code_article]['photos'])
            }
            gestionnaire.articles[code_article]['photos'].append(photo_data)
    
    conn.close()
    return gestionnaire

def sauvegarder_article(code, libelle, emplacement, date_creation):
    """Sauvegarde un article dans SQLite"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO articles (code, libelle, emplacement, date_creation) VALUES (?, ?, ?, ?)",
              (code, libelle, emplacement, date_creation))
    conn.commit()
    conn.close()

def sauvegarder_photo(code_article, timestamp, nb_pieces, algorithme, image_originale, image_analyse):
    """Sauvegarde une photo dans SQLite"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    c.execute("INSERT INTO photos (code_article, timestamp, nb_pieces, algorithme, image_originale, image_analyse) VALUES (?, ?, ?, ?, ?, ?)",
              (code_article, timestamp, nb_pieces, algorithme, image_originale, image_analyse))
    conn.commit()
    conn.close()

def supprimer_article_db(code):
    """Supprime un article de la base"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    c.execute("DELETE FROM photos WHERE code_article = ?", (code,))
    c.execute("DELETE FROM articles WHERE code = ?", (code,))
    conn.commit()
    conn.close()

def supprimer_photo_db(photo_id):
    """Supprime une photo de la base"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    c.execute("DELETE FROM photos WHERE id = ?", (photo_id,))
    conn.commit()
    conn.close()

def get_photo_db_id(code_article, timestamp):
    """Récupère l'ID SQLite d'une photo"""
    conn = sqlite3.connect('inventaire.db')
    c = conn.cursor()
    c.execute("SELECT id FROM photos WHERE code_article = ? AND timestamp = ?", (code_article, timestamp))
    result = c.fetchone()
    conn.close()
    return result[0] if result else None

# ==================== JavaScript pour confirmation ====================

def add_refresh_confirmation():
    has_data = 'true' if 'gestionnaire' in st.session_state and len(st.session_state.gestionnaire.articles) > 0 else 'false'
    refresh_html = f"""
    <div id="refresh-confirmation" style="display:none;"></div>
    <script>
    function hasData() {{
        return {has_data};
    }}
    window.addEventListener('beforeunload', function (e) {{
        if (hasData()) {{
            var confirmationMessage = '⚠️ Attention ! Si vous actualisez la page, toutes les données non exportées seront perdues.\\n\\nVoulez-vous vraiment continuer ?';
            e.returnValue = confirmationMessage;
            return confirmationMessage;
        }}
    }});
    document.addEventListener('keydown', function(e) {{
        if (hasData()) {{
            if (e.key === 'F5' || (e.ctrlKey && e.key === 'r') || (e.ctrlKey && e.key === 'R')) {{
                e.preventDefault();
                var confirmRefresh = confirm('⚠️ Attention ! Si vous actualisez la page, toutes les données non exportées seront perdues.\\n\\nVoulez-vous vraiment actualiser ?');
                if (confirmRefresh) {{
                    window.location.reload();
                }}
            }}
        }}
    }});
    </script>
    """
    st.components.v1.html(refresh_html, height=0)

# ==================== CSS personnalisé ====================

st.markdown("""
<style>
    .success-box {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #28a745;
        margin: 1rem 0;
    }
    .location-badge {
        background: #17a2b8;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 5px;
        font-size: 0.8rem;
        margin-left: 0.5rem;
    }
    .label-badge {
        background: #28a745;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 5px;
        font-size: 0.8rem;
        margin-left: 0.5rem;
    }
    .import-section {
        background: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        border: 2px dashed #6c757d;
        margin: 1rem 0;
    }
    .warning-box {
        background: #fff3cd;
        color: #856404;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #ffc107;
        margin: 1rem 0;
        font-weight: bold;
    }
    .database-info {
        background: #d1ecf1;
        color: #0c5460;
        padding: 0.5rem;
        border-radius: 5px;
        border-left: 5px solid #17a2b8;
        margin: 0.5rem 0;
        font-size: 0.9rem;
    }
    .algorithm-badge {
        background: #6f42c1;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 5px;
        font-size: 0.7rem;
        margin-left: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ==================== Classe GestionnairePieces ====================

class GestionnairePieces:
    def __init__(self):
        """Initialise le gestionnaire de pièces"""
        self.articles = {}
        self.reset_article_courant()
    
    def reset_article_courant(self):
        """Réinitialise l'article en cours de saisie"""
        self.article_courant = {
            'code': '',
            'libelle': '',
            'emplacement': '',
            'photos': [],
            'total_pieces': 0
        }
    
    def creer_nouvel_article(self, code_article, libelle="", emplacement=""):
        """Crée un nouvel article dans l'inventaire"""
        if code_article and code_article not in self.articles:
            date_creation = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.articles[code_article] = {
                'libelle': libelle,
                'photos': [],
                'emplacement': emplacement,
                'date_creation': date_creation
            }
            sauvegarder_article(code_article, libelle, emplacement, date_creation)
            return True
        return False
    
    def nettoyer_articles_mal_importes(self):
        """Supprime les articles qui ont des libellés d'en-tête"""
        a_supprimer = []
        for code, data in self.articles.items():
            libelle = data.get('libelle', '').upper()
            if 'COLONNE' in libelle or 'CODE ARTICLE' in libelle or 'LIBELLÉ' in libelle or 'EMPLACEMENT' in libelle:
                a_supprimer.append(code)
        
        for code in a_supprimer:
            supprimer_article_db(code)
            del self.articles[code]
        
        return len(a_supprimer)
    
    def importer_articles_excel(self, df, col_code, col_libelle, col_emplacement, skip_first_row=True):
        """Importe des articles depuis Excel"""
        articles_importes = 0
        articles_existants = 0
        erreurs = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        start_idx = 1 if skip_first_row else 0
        total_lignes = len(df) - start_idx
        codes_vus = set()
        
        for index in range(start_idx, len(df)):
            progression = (index - start_idx + 1) / total_lignes
            progress_bar.progress(progression)
            status_text.text(f"Import en cours... {index - start_idx + 1}/{total_lignes}")
            
            row = df.iloc[index]
            try:
                code_value = row[col_code]
                if pd.isna(code_value) or str(code_value).strip() == '':
                    continue
                
                code = str(code_value).strip()
                
                if code.lower() in ['code article', 'code', 'article', 'réf', 'ref', '0', '1', '2', '3', '4']:
                    continue
                
                if code in codes_vus:
                    continue
                codes_vus.add(code)
                
                libelle = ""
                if col_libelle and col_libelle != "(Aucune)" and col_libelle in row.index:
                    libelle_value = row[col_libelle]
                    if pd.notna(libelle_value):
                        libelle = str(libelle_value).strip()
                        if libelle.lower() == 'none':
                            libelle = ""
                
                emplacement = ""
                if col_emplacement and col_emplacement != "(Aucune)" and col_emplacement in row.index:
                    emp_value = row[col_emplacement]
                    if pd.notna(emp_value):
                        emp_str = str(emp_value).strip()
                        if emp_str.lower() not in ['none', 'nan', '']:
                            emplacement = emp_str
                
                if code and code not in self.articles:
                    date_creation = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.articles[code] = {
                        'libelle': libelle,
                        'photos': [],
                        'emplacement': emplacement,
                        'date_creation': date_creation
                    }
                    sauvegarder_article(code, libelle, emplacement, date_creation)
                    articles_importes += 1
                elif code in self.articles:
                    articles_existants += 1
                        
            except Exception as e:
                erreurs += 1
                continue
        
        progress_bar.empty()
        status_text.empty()
        
        return articles_importes, articles_existants, erreurs
    
    def ajouter_photo_article(self, code_article, frame_original, frame_analyse, nb_pieces, algorithme):
        """Ajoute une photo analysée à un article"""
        if code_article in self.articles:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            _, buffer_original = cv2.imencode('.jpg', frame_original)
            _, buffer_analyse = cv2.imencode('.jpg', frame_analyse)
            
            img_originale_b64 = base64.b64encode(buffer_original).decode('utf-8')
            img_analyse_b64 = base64.b64encode(buffer_analyse).decode('utf-8')
            
            photo_data = {
                'timestamp': timestamp,
                'nb_pieces': nb_pieces,
                'algorithme': algorithme,
                'image_originale': img_originale_b64,
                'image_analyse': img_analyse_b64,
                'id': len(self.articles[code_article]['photos'])
            }
            
            self.articles[code_article]['photos'].append(photo_data)
            sauvegarder_photo(code_article, timestamp, nb_pieces, algorithme, img_originale_b64, img_analyse_b64)
            
            return True
        return False
    
    def get_total_article(self, code_article):
        """Retourne le total de pièces pour un article"""
        if code_article in self.articles:
            return sum(photo['nb_pieces'] for photo in self.articles[code_article]['photos'])
        return 0
    
    def get_photos_article(self, code_article):
        """Retourne toutes les photos d'un article"""
        if code_article in self.articles:
            return self.articles[code_article]['photos']
        return []
    
    def get_emplacement_article(self, code_article):
        """Retourne l'emplacement d'un article"""
        if code_article in self.articles:
            return self.articles[code_article].get('emplacement', '')
        return ''
    
    def get_libelle_article(self, code_article):
        """Retourne le libellé d'un article"""
        if code_article in self.articles:
            return self.articles[code_article].get('libelle', '')
        return ''
    
    def supprimer_photo(self, code_article, photo_id):
        """Supprime une photo d'un article"""
        if code_article in self.articles and 0 <= photo_id < len(self.articles[code_article]['photos']):
            timestamp = self.articles[code_article]['photos'][photo_id]['timestamp']
            db_id = get_photo_db_id(code_article, timestamp)
            
            if db_id:
                supprimer_photo_db(db_id)
            
            del self.articles[code_article]['photos'][photo_id]
            
            for i, photo in enumerate(self.articles[code_article]['photos']):
                photo['id'] = i
            return True
        return False
    
    def supprimer_article(self, code_article):
        """Supprime complètement un article"""
        if code_article in self.articles:
            supprimer_article_db(code_article)
            del self.articles[code_article]
            return True
        return False
    
    def get_tous_les_totaux(self):
        """Retourne tous les totaux par article"""
        return {code: self.get_total_article(code) for code in self.articles}
    
    def get_tous_emplacements(self):
        """Retourne tous les emplacements par article"""
        return {code: self.get_emplacement_article(code) for code in self.articles}
    
    def get_tous_libelles(self):
        """Retourne tous les libellés par article"""
        return {code: self.get_libelle_article(code) for code in self.articles}
    
    def generer_excel(self):
        """Génère un fichier Excel avec l'inventaire complet"""
        output = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Feuille principale
        sheet_resume = workbook.active
        sheet_resume.title = "Inventaire"
        
        headers = ["Code Article", "Libellé", "Emplacement", "Quantité totale", "Nombre de photos", "Dernière mise à jour"]
        for col, header in enumerate(headers, 1):
            cell = sheet_resume.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for code_article, data in self.articles.items():
            total = sum(p['nb_pieces'] for p in data['photos'])
            nb_photos = len(data['photos'])
            derniere_date = data['photos'][-1]['timestamp'] if data['photos'] else data.get('date_creation', 'N/A')
            emplacement = data.get('emplacement', '')
            libelle = data.get('libelle', '')
            
            sheet_resume.cell(row=row, column=1).value = code_article
            sheet_resume.cell(row=row, column=2).value = libelle
            sheet_resume.cell(row=row, column=3).value = emplacement
            sheet_resume.cell(row=row, column=4).value = total
            sheet_resume.cell(row=row, column=5).value = nb_photos
            sheet_resume.cell(row=row, column=6).value = derniere_date
            row += 1
        
        # Feuille de détail
        sheet_detail = workbook.create_sheet("Détail des photos")
        
        detail_headers = ["Code Article", "Libellé", "Emplacement", "Photo #", "Date", "Algorithme", "Nombre de pièces"]
        for col, header in enumerate(detail_headers, 1):
            cell = sheet_detail.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for code_article, data in self.articles.items():
            libelle = data.get('libelle', '')
            emplacement = data.get('emplacement', '')
            for i, photo in enumerate(data['photos'], 1):
                sheet_detail.cell(row=row, column=1).value = code_article
                sheet_detail.cell(row=row, column=2).value = libelle
                sheet_detail.cell(row=row, column=3).value = emplacement
                sheet_detail.cell(row=row, column=4).value = f"Photo {i}"
                sheet_detail.cell(row=row, column=5).value = photo['timestamp']
                sheet_detail.cell(row=row, column=6).value = photo.get('algorithme', 'unknown')
                sheet_detail.cell(row=row, column=7).value = photo['nb_pieces']
                row += 1
        
        workbook.save(output)
        output.seek(0)
        return output
    
    def reinitialiser_tout(self):
        """Réinitialise complètement l'inventaire"""
        if os.path.exists('inventaire.db'):
            os.remove('inventaire.db')
        self.articles = {}

# ==================== Fonctions utilitaires ====================

def recadrer_selon_ratio(image, ratio):
    """Recadre l'image au centre pour obtenir le ratio spécifié"""
    if ratio is None:
        return image
    h, w = image.shape[:2]
    ratio_actuel = w / h
    if abs(ratio_actuel - ratio) < 0.01:
        return image
    if ratio_actuel > ratio:
        nouvelle_largeur = int(h * ratio)
        debut_x = (w - nouvelle_largeur) // 2
        return image[:, debut_x:debut_x+nouvelle_largeur]
    else:
        nouvelle_hauteur = int(w / ratio)
        debut_y = (h - nouvelle_hauteur) // 2
        return image[debut_y:debut_y+nouvelle_hauteur, :]

def base64_to_image(base64_string):
    """Convertit une chaîne base64 en image"""
    img_data = base64.b64decode(base64_string)
    nparr = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    return img

# ==================== Initialisation ====================

init_database()

if 'gestionnaire' not in st.session_state:
    st.session_state.gestionnaire = charger_donnees()
if 'page' not in st.session_state:
    st.session_state.page = "saisie"
if 'article_selectionne' not in st.session_state:
    st.session_state.article_selectionne = None
if 'photo_selectionnee' not in st.session_state:
    st.session_state.photo_selectionnee = None
if 'show_import' not in st.session_state:
    st.session_state.show_import = False
if 'photo_temp' not in st.session_state:
    st.session_state.photo_temp = None
if 'ajout_photo' not in st.session_state:
    st.session_state.ajout_photo = False
if 'search_query' not in st.session_state:
    st.session_state.search_query = ""

gestionnaire = st.session_state.gestionnaire

add_refresh_confirmation()

# ==================== Affichage des avertissements ====================

if len(gestionnaire.articles) > 0:
    st.markdown("""
    <div class="warning-box">
        ⚠️ <strong>Attention :</strong> Les données sont sauvegardées automatiquement.
        Pensez à exporter votre inventaire en Excel pour une sauvegarde externe !
    </div>
    """, unsafe_allow_html=True)

st.markdown("""
<div class="database-info">
    💾 <strong>Persistance active :</strong> Les données sont automatiquement sauvegardées dans une base SQLite
</div>
""", unsafe_allow_html=True)

# ==================== Interface principale ====================

st.title("📦 Gestionnaire d'Inventaire Multi-Pièces")
st.markdown("""
Cette application permet de gérer l'inventaire de plusieurs types de pièces :
1. **Importer** un fichier Excel avec vos articles (code, libellé, emplacement)
2. **Ajouter** plusieurs photos pour chaque article (avec choix de l'algorithme de détection)
3. **Exporter** un fichier Excel avec tous les totaux
""")

# ==================== Barre latérale ====================

with st.sidebar:
    st.header("📋 Articles en inventaire")
    
    if st.button("📥 Importer des articles Excel", use_container_width=True):
        st.session_state.show_import = True
        st.rerun()
    
    if gestionnaire.articles:
        st.write(f"**{len(gestionnaire.articles)} articles**")
        
        if st.button("🧹 Nettoyer les articles mal importés", use_container_width=True):
            nb_supprimes = gestionnaire.nettoyer_articles_mal_importes()
            if nb_supprimes > 0:
                st.success(f"✅ {nb_supprimes} articles supprimés")
                st.rerun()
            else:
                st.info("Aucun article à nettoyer")
        
        search_query = st.text_input(
            "🔍 Rechercher un article",
            value=st.session_state.search_query,
            placeholder="Code, libellé ou emplacement...",
            key="search_input"
        ).lower().strip()
        st.session_state.search_query = search_query
        
        codes_filtres = []
        if search_query:
            for code, data in gestionnaire.articles.items():
                libelle = data.get('libelle', '').lower()
                emplacement = data.get('emplacement', '').lower()
                if (search_query in code.lower() or 
                    search_query in libelle or 
                    search_query in emplacement):
                    codes_filtres.append(code)
        else:
            codes_filtres = list(gestionnaire.articles.keys())
        
        codes_filtres.sort()
        
        if not codes_filtres:
            st.info("Aucun article ne correspond à votre recherche")
        
        for code_article in codes_filtres:
            total = gestionnaire.get_total_article(code_article)
            libelle = gestionnaire.get_libelle_article(code_article)
            emplacement = gestionnaire.get_emplacement_article(code_article)
            
            with st.container():
                col1, col2 = st.columns([3, 1])
                with col1:
                    if st.button(f"📦 {code_article}", key=f"select_{code_article}", use_container_width=True):
                        st.session_state.article_selectionne = code_article
                        st.session_state.page = "details"
                        st.rerun()
                with col2:
                    st.write(f"**{total}**")
            
            if libelle or emplacement:
                badge_text = ""
                if libelle:
                    badge_text += f"📝 {libelle[:30]}{'...' if len(libelle) > 30 else ''}"
                if libelle and emplacement:
                    badge_text += " | "
                if emplacement:
                    badge_text += f"📍 {emplacement}"
                
                if badge_text:
                    st.caption(badge_text)
        
        st.divider()
        
        if st.button("➕ Nouvel article manuel", use_container_width=True):
            st.session_state.page = "saisie"
            st.session_state.article_selectionne = None
            st.rerun()
        
        st.divider()
        
        if gestionnaire.articles:
            st.header("📊 Export")
            excel_file = gestionnaire.generer_excel()
            st.download_button(
                label="📥 Télécharger Excel",
                data=excel_file,
                file_name=f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            if st.button("🔄 Tout réinitialiser", type="primary", use_container_width=True):
                gestionnaire.reinitialiser_tout()
                st.session_state.page = "saisie"
                st.session_state.article_selectionne = None
                st.rerun()
    else:
        st.info("Aucun article pour le moment")

# ==================== Section d'import Excel ====================

if st.session_state.show_import:
    st.markdown("---")
    st.markdown('<div class="import-section">', unsafe_allow_html=True)
    st.header("📥 Importer des articles depuis Excel")
    
    uploaded_excel = st.file_uploader("Choisir un fichier Excel", type=['xlsx', 'xls'], key="import_excel")
    
    if uploaded_excel:
        try:
            df = pd.read_excel(uploaded_excel)
            
            st.subheader("Aperçu du fichier original")
            st.dataframe(df.head(10))
            
            cols = df.columns.tolist()
            
            default_code_index = 0
            default_libelle_index = 1
            default_emplacement_index = 2
            
            for i, col in enumerate(cols):
                col_lower = col.lower()
                if 'emplacement' in col_lower:
                    default_emplacement_index = i + 1
            
            col1, col2, col3 = st.columns(3)
            with col1:
                col_code = st.selectbox("📌 Colonne pour CODE article *", cols, index=default_code_index)
            with col2:
                col_libelle = st.selectbox("📝 Colonne pour LIBELLÉ *", ["(Aucune)"] + cols, index=default_libelle_index + 1)
            with col3:
                col_emplacement = st.selectbox("📍 Colonne pour EMPLACEMENT (optionnel)", ["(Aucune)"] + cols, index=default_emplacement_index)
            
            skip_first = st.checkbox("Ignorer la première ligne (en-têtes)", value=True)
            
            st.subheader("Aperçu des données à importer :")
            
            start_preview = 1 if skip_first else 0
            preview_data = {}
            
            preview_data['Code'] = df[col_code].iloc[start_preview:].values
            
            if col_libelle != "(Aucune)":
                preview_data['Libellé'] = df[col_libelle].iloc[start_preview:].values
            
            if col_emplacement != "(Aucune)":
                preview_data['Emplacement'] = df[col_emplacement].iloc[start_preview:].values
            
            apercu = pd.DataFrame(preview_data)
            st.dataframe(apercu)
            
            total_lignes = len(df) - (1 if skip_first else 0)
            codes_non_vides = df[col_code].iloc[start_preview:].notna().sum()
            codes_uniques = df[col_code].iloc[start_preview:].nunique()
            
            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            with col_s1:
                st.metric("📊 Lignes totales", total_lignes)
            with col_s2:
                st.metric("✅ Codes valides", codes_non_vides)
            with col_s3:
                st.metric("🆔 Codes uniques", codes_uniques)
            with col_s4:
                st.metric("📝 Articles actuels", len(gestionnaire.articles))
            
            if st.button("✅ Confirmer l'import", use_container_width=True, type="primary"):
                with st.spinner("Import en cours..."):
                    col_lib = col_libelle if col_libelle != "(Aucune)" else None
                    col_emp = col_emplacement if col_emplacement != "(Aucune)" else None
                    
                    importes, existants, erreurs = gestionnaire.importer_articles_excel(df, col_code, col_lib, col_emp, skip_first)
                    
                    st.markdown("---")
                    st.subheader("📊 Résultat de l'import")
                    
                    col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                    with col_r1:
                        st.metric("✅ Importés", importes)
                    with col_r2:
                        st.metric("⚠️ Déjà existants", existants)
                    with col_r3:
                        st.metric("❌ Erreurs", erreurs)
                    with col_r4:
                        st.metric("📊 Total après import", len(gestionnaire.articles))
                    
                    if importes > 0:
                        st.success(f"✅ {importes} articles importés avec succès !")
                        st.balloons()
                        st.session_state.show_import = False
                        st.rerun()
                    else:
                        st.warning("⚠️ Aucun article n'a été importé.")
        
        except Exception as e:
            st.error(f"❌ Erreur lors de la lecture du fichier : {str(e)}")
    
    if st.button("❌ Fermer l'import", use_container_width=True):
        st.session_state.show_import = False
        st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")

# ==================== Page de saisie ====================

if st.session_state.page == "saisie" and not st.session_state.show_import:
    st.header("➕ Ajouter un nouvel article")
    
    st.markdown("### 📝 Informations de l'article")
    
    col_code, col_lib, col_emp = st.columns([2, 2, 1])
    
    with col_code:
        code_article = st.text_input(
            "Code article *",
            placeholder="Code article (obligatoire)",
            key="code_article_input"
        )
    
    with col_lib:
        libelle = st.text_input(
            "Libellé (optionnel)",
            value="",
            placeholder="Description de l'article",
            key="libelle_input"
        )
    
    with col_emp:
        emplacement = st.text_input(
            "Emplacement (optionnel)",
            value="",
            placeholder="Ex: A-12, Rayon 3...",
            key="emplacement_input"
        )
    
    st.caption("* Champ obligatoire")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ Créer l'article", use_container_width=True):
            if code_article:
                if gestionnaire.creer_nouvel_article(code_article, libelle, emplacement):
                    st.success(f"✅ Article '{code_article}' créé avec succès!")
                    if libelle:
                        st.info(f"📝 Libellé: {libelle}")
                    if emplacement:
                        st.info(f"📍 Emplacement: {emplacement}")
                    st.session_state.article_selectionne = code_article
                    st.session_state.page = "details"
                    st.rerun()
                else:
                    if code_article in gestionnaire.articles:
                        st.error("❌ Ce code article existe déjà")
                    else:
                        st.error("❌ Erreur lors de la création de l'article")
            else:
                st.error("❌ Veuillez entrer un code article")
    with col2:
        if st.button("❌ Annuler", use_container_width=True):
            st.rerun()

# ==================== Page des détails article ====================

elif st.session_state.page == "details" and st.session_state.article_selectionne:
    code_article = st.session_state.article_selectionne
    photos = gestionnaire.get_photos_article(code_article)
    total = gestionnaire.get_total_article(code_article)
    libelle = gestionnaire.get_libelle_article(code_article)
    emplacement = gestionnaire.get_emplacement_article(code_article)
    
    col_h1, col_h2, col_h3, col_h4, col_h5 = st.columns([2, 1, 1, 1, 1])
    with col_h1:
        st.header(f"📦 {code_article}")
        if libelle:
            st.markdown(f"<span class='label-badge'>📝 {libelle}</span>", unsafe_allow_html=True)
        if emplacement:
            st.markdown(f"<span class='location-badge'>📍 {emplacement}</span>", unsafe_allow_html=True)
    with col_h2:
        st.metric("Total pièces", total)
    with col_h3:
        st.metric("Photos", len(photos))
    with col_h4:
        if libelle:
            st.metric("Libellé", libelle[:20] + "..." if len(libelle) > 20 else libelle)
    with col_h5:
        if emplacement:
            st.metric("Emplacement", emplacement)
    
    col_o1, col_o2, col_o3 = st.columns(3)
    with col_o1:
        if st.button("⬅️ Retour à la saisie", use_container_width=True):
            st.session_state.page = "saisie"
            st.rerun()
    with col_o2:
        if st.button("📸 Ajouter une photo", use_container_width=True):
            st.session_state.ajout_photo = True
            st.session_state.photo_temp = None
            st.rerun()
    with col_o3:
        if st.button("🗑️ Supprimer cet article", use_container_width=True, type="primary"):
            if gestionnaire.supprimer_article(code_article):
                st.success(f"✅ Article '{code_article}' supprimé")
                st.session_state.page = "saisie"
                st.rerun()
    
    st.divider()
    
    # ==================== Ajout de photo ====================
    
    if st.session_state.get('ajout_photo', False):
        st.subheader("📸 Ajouter une photo")
        
        col_p1, col_p2 = st.columns([2, 1])
        with col_p2:
            if st.button("❌ Annuler"):
                st.session_state.ajout_photo = False
                st.session_state.photo_temp = None
                st.rerun()
        
        with col_p1:
            source = st.radio("Source", ["📸 Prendre une photo", "🖼️ Choisir une image"], horizontal=True, key="photo_source")
        
        img_file = None
        if source == "📸 Prendre une photo":
            img_file = st.camera_input("Prendre une photo", key="camera_photo")
        else:
            img_file = st.file_uploader("Choisir une image", type=['jpg', 'jpeg', 'png'], key="upload_photo")
        
        if img_file is not None:
            with st.spinner("Chargement de l'image..."):
                bytes_data = img_file.getvalue()
                frame_brut = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
                st.session_state.photo_temp = {
                    'brut': frame_brut,
                    'format_choisi': "Original",
                    'recadree': None,
                    'analyse': None,
                    'detected': 0,
                    'algorithme': "watershed"
                }
        
        if st.session_state.photo_temp is not None:
            temp = st.session_state.photo_temp
            frame_brut = temp['brut']
            
            format_options = ["Original", "4:3", "16:9"]
            index_defaut = format_options.index(temp.get('format_choisi', "Original"))
            format_choisi = st.selectbox("Format d'image", format_options, index=index_defaut, key="format_select")
            
            # Choix de l'algorithme
            algo_options = {
                "🌊 Watershed (recommandé - sépare les objets collés)": "watershed",
                "⭕ Cercles Hough (idéal pour pièces rondes)": "hough_cercles",
                "🔷 Contours classiques (rapide - fusionne les objets collés)": "contours"
            }
            algo_choisi = st.selectbox(
                "🔬 Algorithme de détection",
                options=list(algo_options.keys()),
                index=0,
                help="Watershed: Sépare les objets qui se touchent\nCercles Hough: Idéal pour vis, rondelles, pièces de monnaie\nContours: Simple mais fusionne les objets collés"
            )
            algorithme = algo_options[algo_choisi]
            
            if format_choisi != temp.get('format_choisi') or algorithme != temp.get('algorithme'):
                temp['format_choisi'] = format_choisi
                temp['algorithme'] = algorithme
                
                if format_choisi == "4:3":
                    frame_recadree = recadrer_selon_ratio(frame_brut, 4/3)
                elif format_choisi == "16:9":
                    frame_recadree = recadrer_selon_ratio(frame_brut, 16/9)
                else:
                    frame_recadree = frame_brut
                temp['recadree'] = frame_recadree
                
                resultat, nb = detecter_pieces(frame_recadree, algorithme)
                temp['analyse'] = resultat
                temp['detected'] = nb
            
            if temp.get('analyse') is None:
                if format_choisi == "4:3":
                    frame_recadree = recadrer_selon_ratio(frame_brut, 4/3)
                elif format_choisi == "16:9":
                    frame_recadree = recadrer_selon_ratio(frame_brut, 16/9)
                else:
                    frame_recadree = frame_brut
                temp['recadree'] = frame_recadree
                resultat, nb = detecter_pieces(frame_recadree, algorithme)
                temp['analyse'] = resultat
                temp['detected'] = nb
            
            st.image(cv2.cvtColor(temp['analyse'], cv2.COLOR_BGR2RGB), 
                     caption=f"Analyse - {temp['detected']} pièces détectées avec {algorithme}", 
                     use_container_width=True)
            
            st.markdown("### Options de comptage")
            col_opt1, col_opt2, col_opt3 = st.columns(3)
            
            with col_opt1:
                operation = st.selectbox("Opération", 
                                         ["Utiliser détection", "Remplacer", "Additionner", "Multiplier"],
                                         index=0)
            with col_opt2:
                manuel = st.number_input("Valeur manuelle", min_value=0, value=0, step=1)
            with col_opt3:
                st.write("")
                st.write("")
                if st.button("✅ Ajouter cette photo", use_container_width=True):
                    detected = temp['detected']
                    if operation == "Utiliser détection":
                        nb_final = detected
                    elif operation == "Remplacer":
                        nb_final = manuel if manuel > 0 else detected
                    elif operation == "Additionner":
                        nb_final = detected + manuel
                    elif operation == "Multiplier":
                        nb_final = detected * manuel if manuel > 0 else detected
                    else:
                        nb_final = detected
                    
                    if gestionnaire.ajouter_photo_article(code_article, temp['recadree'], temp['analyse'], nb_final, algorithme):
                        st.success(f"✅ Photo ajoutée avec {nb_final} pièces!")
                        st.session_state.ajout_photo = False
                        st.session_state.photo_temp = None
                        st.rerun()
    
    # ==================== Affichage des photos existantes ====================
    
    if photos:
        st.subheader("📸 Photos enregistrées")
        
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            tri = st.selectbox("Trier par", ["Plus récente", "Plus ancienne", "Plus de pièces", "Moins de pièces"])
        
        photos_affichees = photos.copy()
        if tri == "Plus récente":
            photos_affichees = list(reversed(photos_affichees))
        elif tri == "Plus de pièces":
            photos_affichees = sorted(photos_affichees, key=lambda x: x['nb_pieces'], reverse=True)
        elif tri == "Moins de pièces":
            photos_affichees = sorted(photos_affichees, key=lambda x: x['nb_pieces'])
        
        cols = st.columns(3)
        for i, photo in enumerate(photos_affichees):
            with cols[i % 3]:
                img = base64_to_image(photo['image_analyse'])
                img_mini = cv2.resize(img, (200, 150))
                st.image(cv2.cvtColor(img_mini, cv2.COLOR_BGR2RGB), use_column_width=True)
                
                st.caption(f"📅 {photo['timestamp'][:10]}")
                st.caption(f"🔢 {photo['nb_pieces']} pièces")
                if 'algorithme' in photo:
                    st.caption(f"🔬 {photo['algorithme']}")
                
                col_b1, col_b2 = st.columns(2)
                with col_b1:
                    if st.button("🔍 Voir", key=f"view_{code_article}_{i}"):
                        st.session_state.photo_selectionnee = photo['id']
                        st.session_state.page = "photo_detail"
                        st.rerun()
                with col_b2:
                    if st.button("🗑️", key=f"del_{code_article}_{i}"):
                        if gestionnaire.supprimer_photo(code_article, photo['id']):
                            st.rerun()
    else:
        st.info("📸 Aucune photo pour cet article. Cliquez sur 'Ajouter une photo' pour commencer.")

# ==================== Page de détail photo ====================

elif st.session_state.page == "photo_detail" and st.session_state.article_selectionne and st.session_state.photo_selectionnee is not None:
    code_article = st.session_state.article_selectionne
    photos = gestionnaire.get_photos_article(code_article)
    photo_id = st.session_state.photo_selectionnee
    
    if 0 <= photo_id < len(photos):
        photo = photos[photo_id]
        libelle = gestionnaire.get_libelle_article(code_article)
        
        st.header(f"🔍 Détail de la photo - {code_article}")
        if libelle:
            st.subheader(libelle)
        
        col_img1, col_img2 = st.columns(2)
        
        with col_img1:
            st.subheader("📸 Image originale")
            img_originale = base64_to_image(photo['image_originale'])
            st.image(cv2.cvtColor(img_originale, cv2.COLOR_BGR2RGB), use_column_width=True)
        
        with col_img2:
            st.subheader(f"🔍 Analyse - {photo['nb_pieces']} pièces")
            if 'algorithme' in photo:
                st.markdown(f"<span class='algorithm-badge'>🔬 {photo['algorithme']}</span>", unsafe_allow_html=True)
            img_analyse = base64_to_image(photo['image_analyse'])
            st.image(cv2.cvtColor(img_analyse, cv2.COLOR_BGR2RGB), use_column_width=True)
        
        st.metric("Nombre de pièces", photo['nb_pieces'])
        st.caption(f"Date: {photo['timestamp']}")
        
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            if st.button("⬅️ Retour à l'article", use_container_width=True):
                st.session_state.page = "details"
                st.session_state.photo_selectionnee = None
                st.rerun()
        with col_b2:
            if st.button("🗑️ Supprimer cette photo", use_container_width=True, type="primary"):
                if gestionnaire.supprimer_photo(code_article, photo_id):
                    st.session_state.page = "details"
                    st.session_state.photo_selectionnee = None
                    st.rerun()
    else:
        st.error("Photo non trouvée")
        if st.button("Retour"):
            st.session_state.page = "details"
            st.session_state.photo_selectionnee = None
            st.rerun()

# ==================== Pied de page ====================

st.markdown("---")
col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
with col_f1:
    st.caption("📦 Gestionnaire d'Inventaire")
with col_f2:
    total_global = sum(gestionnaire.get_tous_les_totaux().values())
    st.caption(f"🧩 Total global: {total_global} pièces")
with col_f3:
    st.caption(f"📊 Articles: {len(gestionnaire.articles)}")
with col_f4:
    emplacements_renseignes = sum(1 for e in gestionnaire.get_tous_emplacements().values() if e)
    st.caption(f"📍 Emplacements: {emplacements_renseignes}/{len(gestionnaire.articles)}")
with col_f5:
    libelles_renseignes = sum(1 for l in gestionnaire.get_tous_libelles().values() if l)
    st.caption(f"📝 Libellés: {libelles_renseignes}/{len(gestionnaire.articles)}")
